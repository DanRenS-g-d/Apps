import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Configuration
INPUT_FOLDER = "scrapers_output"
OUTPUT_FILE = "unified_supermarket_data.xlsx"
PRIMARY_KEY = "Título"  # Main product identifier
STORE_COL = "Tienda"  # Column to identify source store
DATE_COL = "Fecha_Extracción"  # Column for scrape date

# Store name mappings from filenames
STORE_MAPPINGS = {
    'd1': 'D1',
    'ara': 'Ara',
    'isimo': 'Ísimo',
    'olimpica': 'Olimpica',
    'exito': 'Éxito'
}

def unify_supermarket_data():
    start_time = datetime.now()
    print(f"\n🚀 Starting unification at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    all_data = []
    file_stats = []

    # 1. Process all Excel files
    print("\n📂 Processing files...")
    for filename in os.listdir(INPUT_FOLDER):
        if filename.lower().endswith(('.xlsx', '.xls')):
            filepath = os.path.join(INPUT_FOLDER, filename)
            try:
                store_name = identify_store(filename)
                print(f"  🔍 Found {store_name} data in {filename}")
                
                # Read all sheets from Excel file
                xls = pd.ExcelFile(filepath)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    
                    # Standardize column names
                    df.columns = standardize_columns(df.columns)
                    
                    # Add metadata
                    df[STORE_COL] = store_name
                    df[DATE_COL] = datetime.now().date()
                    
                    # Clean price data
                    df = clean_price_data(df)
                    
                    all_data.append(df)
                    file_stats.append({
                        'File': filename,
                        'Store': store_name,
                        'Sheet': sheet_name,
                        'Records': len(df)
                    })
                    
            except Exception as e:
                print(f"  ❌ Error processing {filename}: {str(e)}")
                continue

    if not all_data:
        print("\n⚠️ No valid data files found!")
        return

    # 2. Combine all data
    print("\n🔗 Merging data...")
    unified_df = pd.concat(all_data, ignore_index=True)
    initial_count = len(unified_df)

    # 3. Data cleaning
    print("\n🧹 Cleaning data...")
    
    # Text standardization
    text_cols = unified_df.select_dtypes(include=['object']).columns
    for col in text_cols:
        unified_df[col] = unified_df[col].astype(str).str.strip().str.upper()
    
    # Handle duplicates
    unified_df.drop_duplicates(
        subset=[PRIMARY_KEY, STORE_COL], 
        keep='last', 
        inplace=True
    )
    
    # Final cleanup
    unified_df = unified_df[~unified_df[PRIMARY_KEY].isin(['N/A', 'NAN', 'NONE'])]
    unified_df.dropna(subset=[PRIMARY_KEY], inplace=True)
    
    # 4. Save results
    print("\n💾 Saving unified file...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Main data sheet
        unified_df.to_excel(writer, sheet_name='DATOS_UNIFICADOS', index=False)
        
        # Create analytical sheets
        create_analytical_sheets(writer, unified_df)
        
        # Create metadata sheet
        pd.DataFrame(file_stats).to_excel(
            writer, 
            sheet_name='METADATOS', 
            index=False
        )

    # 5. Format Excel file
    format_excel_output(OUTPUT_FILE)
    
    # 6. Final report
    end_time = datetime.now()
    duration = end_time - start_time
    
    print(f"""
    ╔══════════════════════════════════════╗
    ║         UNIFICATION COMPLETE         ║
    ╠══════════════════════════════════════╣
    ║ Stores Processed: {len(STORE_MAPPINGS):<5}           ║
    ║ Files Processed:  {len(file_stats):<5}           ║
    ║ Initial Records:  {initial_count:<10}      ║
    ║ Final Records:    {len(unified_df):<10}      ║
    ║ Duplicates Removed: {initial_count-len(unified_df):<7}      ║
    ║ Execution Time:   {duration.seconds} seconds       ║
    ╚══════════════════════════════════════╝
    """)
    print(f"📊 Unified file saved as: {OUTPUT_FILE}")

def identify_store(filename):
    """Extract store name from filename"""
    filename_lower = filename.lower()
    for key, value in STORE_MAPPINGS.items():
        if key in filename_lower:
            return value
    # Fallback: extract from filename
    return os.path.splitext(filename)[0].split('_')[-1].upper()

def standardize_columns(columns):
    """Standardize column names across different scrapers"""
    column_mapping = {
        'nombre': 'Título',
        'producto': 'Título',
        'descripción': 'Título',
        'precio': 'Precio',
        'valor': 'Precio',
        'price': 'Precio',
        'categoría': 'Categoría',
        'category': 'Categoría',
        'tipo': 'Categoría',
        'subcategoría': 'Subcategoría',
        'subcategory': 'Subcategoría'
    }
    
    return [column_mapping.get(col.lower().strip(), col) for col in columns]

def clean_price_data(df):
    """Clean and standardize price columns"""
    price_cols = [col for col in df.columns if 'PRECIO' in col.upper() or 'PRICE' in col.upper()]
    
    for col in price_cols:
        if df[col].dtype == 'object':
            # Remove currency symbols and thousands separators
            df[col] = (
                df[col].astype(str)
                .str.replace(r'[^\d.,]', '', regex=True)
                .str.replace('.', '', regex=False)
                .str.replace(',', '.', regex=False)
                .astype(float)
            )
            # Fix prices that are too low (e.g., 2.8 instead of 2800 COP)
        df[col] = df[col].apply(lambda x: x * 1000 if x < 50 else x)
        
    return df

def create_analytical_sheets(writer, df):
    """Create analytical sheets in the Excel file"""
    
    # Price comparison by store
    if 'Precio' in df.columns:
        price_pivot = df.pivot_table(
            index=PRIMARY_KEY,
            columns=STORE_COL,
            values='Precio',
            aggfunc='first'
        )
        price_pivot['MIN_PRICE'] = price_pivot.min(axis=1)
        price_pivot['MAX_PRICE'] = price_pivot.max(axis=1)
        price_pivot['PRICE_DIFF'] = price_pivot['MAX_PRICE'] - price_pivot['MIN_PRICE']
        price_pivot.to_excel(writer, sheet_name='COMPARACION_PRECIOS')

    # Category analysis
    if 'Categoría' in df.columns:
        category_stats = df.groupby([STORE_COL, 'Categoría']).size().unstack(fill_value=0)
        category_stats['TOTAL'] = category_stats.sum(axis=1)
        category_stats.to_excel(writer, sheet_name='ANALISIS_CATEGORIAS')

    # Price distribution analysis
    if 'Precio' in df.columns:
        price_stats = df.groupby(STORE_COL)['Precio'].agg(['count', 'min', 'max', 'mean', 'median'])
        price_stats.to_excel(writer, sheet_name='ESTADISTICAS_PRECIOS')

def format_excel_output(filepath):
    """Apply formatting to the Excel output file"""
    wb = load_workbook(filepath)
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Set column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Freeze panes for easy navigation
    for sheet in wb:
        sheet.freeze_panes = 'A2'
    
    wb.save(filepath)

if __name__ == "__main__":
    # Verify input folder exists
    if not os.path.exists(INPUT_FOLDER):
        os.makedirs(INPUT_FOLDER)
        print(f"\n📁 Created folder '{INPUT_FOLDER}'. Please place your scraper Excel files there.")
    else:
        unify_supermarket_data()